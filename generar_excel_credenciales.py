import csv
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ── Leer CSV ────────────────────────────────────────
with open('credenciales_facilitadores.csv', encoding='utf-8-sig') as f:
    rows = list(csv.DictReader(f))

# ── Workbook ─────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = 'Credenciales Facilitadores'

# ── Colores ──────────────────────────────────────────
VERDE_OSCURO  = '1A5E20'
VERDE_MEDIO   = '2E7D32'
VERDE_CLARO   = 'E8F5E9'
VERDE_ALT     = 'F1F8E9'
BLANCO        = 'FFFFFF'
GRIS_TEXTO    = '37474F'
NARANJA       = 'E65100'

thin = Side(style='thin', color='B0BEC5')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── Título principal (fila 1, columnas A–E combinadas) ──
ws.merge_cells('A1:E1')
title_cell = ws['A1']
title_cell.value = 'SEMBRANDO VIDA — Credenciales Facilitadores Comunitarios'
title_cell.font = Font(name='Calibri', size=16, bold=True, color=BLANCO)
title_cell.fill = PatternFill('solid', fgColor=VERDE_OSCURO)
title_cell.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 36

# ── Subtítulo (fila 2) ────────────────────────────────
ws.merge_cells('A2:E2')
sub = ws['A2']
sub.value = f'Total: {len(rows)} facilitadores  |  URL: https://adminpwa.sembrandodatos.com  |  Generado: 13/04/2026'
sub.font = Font(name='Calibri', size=10, italic=True, color=BLANCO)
sub.fill = PatternFill('solid', fgColor=VERDE_MEDIO)
sub.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[2].height = 20

# ── Aviso (fila 3) ────────────────────────────────────
ws.merge_cells('A3:E3')
aviso = ws['A3']
aviso.value = '⚠  CONFIDENCIAL — Distribuir únicamente al facilitador correspondiente. No compartir con terceros.'
aviso.font = Font(name='Calibri', size=9, bold=True, color=NARANJA)
aviso.fill = PatternFill('solid', fgColor='FFF8E1')
aviso.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[3].height = 18

# ── Encabezados (fila 4) ──────────────────────────────
headers = ['#', 'Nombre Completo', 'CURP', 'Usuario', 'Contraseña']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col, value=h)
    cell.font = Font(name='Calibri', size=11, bold=True, color=BLANCO)
    cell.fill = PatternFill('solid', fgColor=VERDE_MEDIO)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border
ws.row_dimensions[4].height = 22

# ── Datos ─────────────────────────────────────────────
for i, row in enumerate(rows, 1):
    excel_row = i + 4
    fill_color = VERDE_CLARO if i % 2 == 0 else VERDE_ALT
    fill = PatternFill('solid', fgColor=fill_color)

    data = [
        i,
        row.get('nombre', '').upper(),
        row.get('curp', '').upper(),
        row.get('username', ''),
        row.get('password', ''),
    ]
    aligns = ['center', 'left', 'center', 'center', 'center']

    for col, (val, al) in enumerate(zip(data, aligns), 1):
        cell = ws.cell(row=excel_row, column=col, value=val)
        cell.font = Font(name='Calibri', size=10, color=GRIS_TEXTO)
        cell.fill = fill
        cell.alignment = Alignment(horizontal=al, vertical='center')
        cell.border = border
    ws.row_dimensions[excel_row].height = 18

# ── Anchos de columna ─────────────────────────────────
anchos = [5, 46, 22, 16, 18]
for col, ancho in enumerate(anchos, 1):
    ws.column_dimensions[get_column_letter(col)].width = ancho

# ── Congelar encabezados ──────────────────────────────
ws.freeze_panes = 'A5'

# ── Autofilter ────────────────────────────────────────
ws.auto_filter.ref = f'A4:E{len(rows)+4}'

# ── Pie de página ─────────────────────────────────────
pie_row = len(rows) + 5
ws.merge_cells(f'A{pie_row}:E{pie_row}')
pie = ws[f'A{pie_row}']
pie.value = 'Sembrando Vida · Programa de Bienestar · México 2026 · Acceso: https://adminpwa.sembrandodatos.com'
pie.font = Font(name='Calibri', size=9, italic=True, color='78909C')
pie.fill = PatternFill('solid', fgColor='ECEFF1')
pie.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[pie_row].height = 16

# ── Guardar ───────────────────────────────────────────
out = 'credenciales_facilitadores.xlsx'
wb.save(out)
print(f'✅ Generado: {out}  ({len(rows)} filas)')
